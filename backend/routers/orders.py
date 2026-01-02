"""
订单管理路由
"""

from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_, func
from backend.database import get_db
from backend.models import Order, Supplier, User, Product
from backend.schemas import (
    OrderCreate, OrderResponse, OrderDetailResponse, OrderListResponse, OrderFilter, OrderItem
)
from backend.auth import get_current_user, require_admin, can_view_internal_price
from backend.utils import format_order_content, parse_order_content, calculate_order_totals, remove_internal_price_from_items
from backend.config import (
    USER_TYPE_ADMIN, USER_TYPE_NORMAL, USER_TYPE_SUPPLIER, USER_TYPE_STUDENT,
    ORDER_STATUS_DRAFT, ORDER_STATUS_SUBMITTED, ORDER_STATUS_CONFIRMED, ORDER_STATUS_INVALID
)
from backend.logger import get_logger
from backend.email_sender import send_order_notification
from typing import Optional
from datetime import datetime
import json
import io
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

logger = get_logger(__name__)

router = APIRouter(prefix="/api/orders", tags=["orders"])


@router.get("/", response_model=OrderListResponse)
async def list_orders(
    supplier_id: Optional[int] = Query(None, description="厂家ID筛选"),
    content: Optional[str] = Query(None, description="订单内容筛选"),
    min_amount: Optional[float] = Query(None, ge=0, description="最低金额"),
    max_amount: Optional[float] = Query(None, ge=0, description="最高金额"),
    start_date: Optional[str] = Query(None, description="开始日期"),
    end_date: Optional[str] = Query(None, description="结束日期"),
    status: Optional[str] = Query(None, description="订单状态"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    获取订单列表（支持筛选和分页）
    
    Args:
        supplier_id: 厂家ID筛选
        content: 订单内容筛选
        min_amount: 最低金额
        max_amount: 最高金额
        start_date: 开始日期
        end_date: 结束日期
        status: 订单状态
        page: 页码
        page_size: 每页数量
        current_user: 当前登录用户
        db: 数据库会话
        
    Returns:
        OrderListResponse: 订单列表响应
        
    使用样例:
        GET /api/orders/?page=1&page_size=20
    """
    # 构建查询
    query = db.query(Order)
    
    # 权限控制
    if current_user.user_type == USER_TYPE_ADMIN:
        # 管理员可以看到所有订单
        pass
    elif current_user.user_type == USER_TYPE_NORMAL:
        # 普通用户可以看到自己的订单以及其管理学生的订单
        from sqlalchemy import or_
        # 查询管理的学生用户ID列表
        managed_students = db.query(User.id).filter(User.manager_id == current_user.id).all()
        managed_student_ids = [s[0] for s in managed_students]
        # 自己的订单 + 管理学生的订单
        if managed_student_ids:
            query = query.filter(
                or_(
                    Order.user_id == current_user.id,
                    Order.user_id.in_(managed_student_ids)
                )
            )
        else:
            query = query.filter(Order.user_id == current_user.id)
    elif current_user.user_type == USER_TYPE_STUDENT:
        # 学生用户只能看到自己的订单
        query = query.filter(Order.user_id == current_user.id)
    elif current_user.user_type == USER_TYPE_SUPPLIER:
        # 厂家用户只能看到自己的订单，且看不到暂存状态的订单
        if not current_user.supplier_id:
            # 如果厂家用户没有关联的supplier_id，返回空结果
            query = query.filter(Order.id == -1)  # 永远不匹配的条件
        else:
            query = query.filter(
                and_(
                    Order.supplier_id == current_user.supplier_id,
                    Order.status != ORDER_STATUS_DRAFT
                )
            )
    
    # 筛选条件
    if supplier_id:
        query = query.filter(Order.supplier_id == supplier_id)
    if status:
        query = query.filter(Order.status == status)
    if start_date:
        try:
            start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            query = query.filter(Order.created_at >= start_dt)
        except:
            pass
    if end_date:
        try:
            end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            query = query.filter(Order.created_at <= end_dt)
        except:
            pass
    
    # 获取总数
    total = query.count()
    
    # 分页
    offset = (page - 1) * page_size
    orders = query.order_by(Order.created_at.desc()).offset(offset).limit(page_size).all()
    
    # 加载关联的厂家信息
    for order in orders:
        if not order.supplier:
            order.supplier = db.query(Supplier).filter(Supplier.id == order.supplier_id).first()
    
    # 内容筛选（需要解析JSON）
    if content:
        filtered_orders = []
        for order in orders:
            items = parse_order_content(order.content)
            content_str = json.dumps(items, ensure_ascii=False)
            if content.lower() in content_str.lower():
                filtered_orders.append(order)
        orders = filtered_orders
        total = len(orders)
    
    # 金额筛选（需要计算每个订单的总金额）
    if min_amount is not None or max_amount is not None:
        filtered_orders = []
        for order in orders:
            items = parse_order_content(order.content)
            totals = calculate_order_totals(items, include_internal=False)
            amount = totals["total_tax_included_price"]
            if min_amount is not None and amount < min_amount:
                continue
            if max_amount is not None and amount > max_amount:
                continue
            filtered_orders.append(order)
        orders = filtered_orders
        total = len(orders)
    
    items = []
    for order in orders:
        supplier_name = order.supplier.name if order.supplier else None
        user = db.query(User).filter(User.id == order.user_id).first()
        username = user.username if user else None
        items.append(OrderResponse(
            id=order.id,
            user_id=order.user_id,
            username=username,
            supplier_id=order.supplier_id,
            supplier_name=supplier_name,
            content=order.content,
            status=order.status,
            created_at=order.created_at
        ))
    
    return OrderListResponse(
        total=total,
        items=items,
        page=page,
        page_size=page_size
    )


@router.get("/{order_id}", response_model=OrderDetailResponse)
async def get_order_detail(
    order_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    获取订单详情
    
    Args:
        order_id: 订单ID
        current_user: 当前登录用户
        db: 数据库会话
        
    Returns:
        OrderDetailResponse: 订单详情
        
    Raises:
        HTTPException: 如果订单不存在或无权访问
        
    使用样例:
        GET /api/orders/1
    """
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="订单不存在"
        )
    
    # 权限控制
    if current_user.user_type == USER_TYPE_NORMAL:
        # 普通用户可以访问自己的订单以及其管理学生的订单
        if order.user_id != current_user.id:
            # 检查是否是管理的学生
            student = db.query(User).filter(
                User.id == order.user_id,
                User.manager_id == current_user.id
            ).first()
            if not student:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="无权访问此订单"
                )
    elif current_user.user_type == USER_TYPE_STUDENT:
        # 学生用户只能访问自己的订单
        if order.user_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="无权访问此订单"
            )
    elif current_user.user_type == USER_TYPE_SUPPLIER:
        if not current_user.supplier_id or order.supplier_id != current_user.supplier_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="无权访问此订单"
            )
    
    # 加载厂家信息
    if not order.supplier:
        order.supplier = db.query(Supplier).filter(Supplier.id == order.supplier_id).first()
    
    # 加载用户信息
    user = db.query(User).filter(User.id == order.user_id).first()
    username = user.username if user else None
    
    # 解析订单内容
    items = parse_order_content(order.content)
    can_view_internal = can_view_internal_price(current_user)
    
    # 厂家用户不能看到内部价格
    if not can_view_internal:
        items = remove_internal_price_from_items(items)
    
    # 计算总价
    totals = calculate_order_totals(items, include_internal=can_view_internal)
    
    return OrderDetailResponse(
        id=order.id,
        user_id=order.user_id,
        username=username,
        supplier_id=order.supplier_id,
        supplier_name=order.supplier.name if order.supplier else None,
        items=[OrderItem(**item) if isinstance(item, dict) else item for item in items],
        status=order.status,
        created_at=order.created_at,
        total_internal_price=totals["total_internal_price"] if can_view_internal else None,
        total_tax_included_price=totals["total_tax_included_price"]
    )


@router.post("/", response_model=OrderResponse)
async def create_order(
    order_data: OrderCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    创建订单（普通用户）
    
    Args:
        order_data: 订单创建信息
        current_user: 当前登录用户
        db: 数据库会话
        
    Returns:
        OrderResponse: 创建的订单信息
        
    Raises:
        HTTPException: 如果用户类型不允许或厂家不存在
        
    使用样例:
        POST /api/orders/
        {
            "supplier_id": 1,
            "items": [
                {
                    "product_id": 1,
                    "name": "商品名",
                    "tax_included_price": 100.0,
                    "quantity": 2
                }
            ]
        }
    """
    # 只有普通用户、学生用户和管理员可以创建订单
    if current_user.user_type == USER_TYPE_SUPPLIER:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="厂家用户不能创建订单"
        )
    
    # 验证厂家是否存在
    supplier = db.query(Supplier).filter(Supplier.id == order_data.supplier_id).first()
    if not supplier:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="厂家不存在"
        )
    
    # 检查并补齐内部价格
    items_dict = []
    for item in order_data.items:
        item_dict = item.model_dump()
        # 如果内部价格为空，从数据库查询
        if item_dict.get('internal_price') is None:
            product = db.query(Product).filter(
                Product.id == item_dict['product_id'],
                Product.supplier_id == order_data.supplier_id
            ).first()
            if product:
                item_dict['internal_price'] = product.internal_price
            else:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"商品ID {item_dict['product_id']} 不存在"
                )
        items_dict.append(item_dict)
    
    # 格式化订单内容
    content = format_order_content(items_dict)
    
    # 创建订单
    new_order = Order(
        user_id=current_user.id,
        supplier_id=order_data.supplier_id,
        content=content,
        status=ORDER_STATUS_DRAFT
    )
    db.add(new_order)
    db.commit()
    db.refresh(new_order)
    new_order.supplier = supplier
    
    return OrderResponse(
        id=new_order.id,
        user_id=new_order.user_id,
        supplier_id=new_order.supplier_id,
        supplier_name=supplier.name,
        content=new_order.content,
        status=new_order.status,
        created_at=new_order.created_at
    )


@router.put("/{order_id}/status")
async def update_order_status(
    order_id: int,
    new_status: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    更新订单状态
    
    Args:
        order_id: 订单ID
        new_status: 新状态
        current_user: 当前登录用户
        db: 数据库会话
        
    Returns:
        dict: 更新结果
        
    Raises:
        HTTPException: 如果订单不存在或状态转换无效
        
    使用样例:
        PUT /api/orders/1/status?new_status=确认
    """
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="订单不存在"
        )
    
    # 权限控制
    if current_user.user_type == USER_TYPE_SUPPLIER:
        # 厂家用户只能确认订单
        if new_status != ORDER_STATUS_CONFIRMED:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="厂家用户只能确认订单"
            )
        if order.status != ORDER_STATUS_SUBMITTED:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="只能确认处于发起状态的订单"
            )
        if not current_user.supplier_id or order.supplier_id != current_user.supplier_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="无权操作此订单"
            )
    elif current_user.user_type == USER_TYPE_NORMAL:
        # 普通用户可以操作自己的订单以及其管理学生的订单
        # 检查是否有权限操作此订单
        has_permission = (order.user_id == current_user.id)
        if not has_permission:
            # 检查是否是管理的学生
            student = db.query(User).filter(
                User.id == order.user_id,
                User.manager_id == current_user.id
            ).first()
            has_permission = (student is not None)
        
        if not has_permission:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="无权操作此订单"
            )
        
        if new_status == ORDER_STATUS_SUBMITTED:
            if order.status != ORDER_STATUS_DRAFT:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="只能发起处于暂存状态的订单"
                )
        elif new_status == ORDER_STATUS_INVALID:
            # 普通用户可以删除订单（转为无效）
            pass
        else:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="普通用户只能发起或删除订单"
            )
    elif current_user.user_type == USER_TYPE_STUDENT:
        # 学生用户只能操作自己的订单
        if order.user_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="无权操作此订单"
            )
        if new_status == ORDER_STATUS_SUBMITTED:
            if order.status != ORDER_STATUS_DRAFT:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="只能发起处于暂存状态的订单"
                )
        elif new_status == ORDER_STATUS_INVALID:
            # 学生用户可以删除订单（转为无效）
            pass
        else:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="学生用户只能发起或删除订单"
            )
    # 管理员可以执行任何状态转换
    
    # 验证状态
    valid_statuses = [ORDER_STATUS_DRAFT, ORDER_STATUS_SUBMITTED, ORDER_STATUS_CONFIRMED, ORDER_STATUS_INVALID]
    if new_status not in valid_statuses:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"无效的状态，必须是: {', '.join(valid_statuses)}"
        )
    
    # 加载关联信息（用于邮件通知，在更新状态前加载）
    if not order.supplier:
        order.supplier = db.query(Supplier).filter(Supplier.id == order.supplier_id).first()
    if not order.user:
        order.user = db.query(User).filter(User.id == order.user_id).first()
    
    order.status = new_status
    db.commit()
    db.refresh(order)
    
    # 发送邮件通知
    if new_status == ORDER_STATUS_SUBMITTED:
        # 发起订单：向厂家用户发送通知
        if order.supplier and order.supplier.user and order.supplier.user.email:
            items = parse_order_content(order.content)
            items_summary = ", ".join([f"{item.get('name', '')} x{item.get('quantity', 1)}" for item in items[:3]])
            if len(items) > 3:
                items_summary += f" 等{len(items)}项"
            await send_order_notification(
                to_email=order.supplier.user.email,
                to_name=order.supplier.user.username,
                order_id=order.id,
                order_status=ORDER_STATUS_SUBMITTED,
                supplier_name=order.supplier.name,
                order_summary=items_summary
            )
    elif new_status == ORDER_STATUS_CONFIRMED:
        # 确认订单：向订单创建用户发送通知
        if order.user and order.user.email:
            items = parse_order_content(order.content)
            items_summary = ", ".join([f"{item.get('name', '')} x{item.get('quantity', 1)}" for item in items[:3]])
            if len(items) > 3:
                items_summary += f" 等{len(items)}项"
            await send_order_notification(
                to_email=order.user.email,
                to_name=order.user.username,
                order_id=order.id,
                order_status=ORDER_STATUS_CONFIRMED,
                supplier_name=order.supplier.name if order.supplier else None,
                order_summary=items_summary
            )
    
    logger.info(
        f"订单状态更新成功: 订单ID={order_id}, 新状态={new_status}, 用户={current_user.username}",
        extra={
            "order_id": order_id,
            "new_status": new_status,
            "user_id": current_user.id,
            "user_type": current_user.user_type
        }
    )
    
    return {"success": True, "message": "订单状态更新成功"}


@router.delete("/{order_id}")
async def delete_order(
    order_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    删除订单（普通用户和管理员）
    暂存和发起状态的订单可以直接删除，确认状态的订单删除后将转移到无效状态
    
    Args:
        order_id: 订单ID
        current_user: 当前登录用户
        db: 数据库会话
        
    Returns:
        dict: 删除结果
        
    Raises:
        HTTPException: 如果订单不存在或无权访问
        
    使用样例:
        DELETE /api/orders/1
    """
    if current_user.user_type == USER_TYPE_SUPPLIER:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="厂家用户不能删除订单"
        )
    
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="订单不存在"
        )
    
    # 权限控制
    if current_user.user_type == USER_TYPE_NORMAL:
        # 普通用户可以删除自己的订单以及其管理学生的订单
        if order.user_id != current_user.id:
            # 检查是否是管理的学生
            student = db.query(User).filter(
                User.id == order.user_id,
                User.manager_id == current_user.id
            ).first()
            if not student:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="无权删除此订单"
                )
    elif current_user.user_type == USER_TYPE_STUDENT:
        # 学生用户只能删除自己的订单
        if order.user_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="无权删除此订单"
            )
    
    # 无效的订单不能删除
    if order.status == ORDER_STATUS_INVALID:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="订单已失效"
        )
    
    # 确认状态的订单转移到无效状态，其他状态直接删除
    if order.status == ORDER_STATUS_CONFIRMED:
        order.status = ORDER_STATUS_INVALID
        db.commit()
        return {"success": True, "message": "订单已标记为无效"}
    else:
        db.delete(order)
        db.commit()
        return {"success": True, "message": "订单删除成功"}


@router.get("/{order_id}/export")
async def export_order_excel(
    order_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    导出订单为Excel文件
    
    Args:
        order_id: 订单ID
        current_user: 当前登录用户
        db: 数据库会话
        
    Returns:
        StreamingResponse: Excel文件流
        
    Raises:
        HTTPException: 如果订单不存在或无权访问
        
    使用样例:
        GET /api/orders/1/export
    """
    try:
        order = db.query(Order).filter(Order.id == order_id).first()
        if not order:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="订单不存在"
            )
        
        # 权限控制
        if current_user.user_type == USER_TYPE_NORMAL:
            # 普通用户可以访问自己的订单以及其管理学生的订单
            if order.user_id != current_user.id:
                # 检查是否是管理的学生
                student = db.query(User).filter(
                    User.id == order.user_id,
                    User.manager_id == current_user.id
                ).first()
                if not student:
                    raise HTTPException(
                        status_code=status.HTTP_403_FORBIDDEN,
                        detail="无权访问此订单"
                    )
        elif current_user.user_type == USER_TYPE_STUDENT:
            # 学生用户只能访问自己的订单
            if order.user_id != current_user.id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="无权访问此订单"
                )
        elif current_user.user_type == USER_TYPE_SUPPLIER:
            if order.supplier_id != current_user.id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="无权访问此订单"
                )
        
        # 加载厂家信息
        if not order.supplier:
            order.supplier = db.query(Supplier).filter(Supplier.id == order.supplier_id).first()
        
        # 解析订单内容
        items = parse_order_content(order.content)
        can_view_internal = can_view_internal_price(current_user)
        
        # 厂家用户不能看到内部价格
        if not can_view_internal:
            items = remove_internal_price_from_items(items)
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = f"订单_{order_id}"
        
        # 设置标题样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入表头
        headers = ["商品名", "品牌", "型号", "规格"]
        if can_view_internal:
            headers.append("内部价格")
        headers.extend(["含税价格", "数量", "小计"])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 写入数据
        for row, item in enumerate(items, 2):
            col = 1
            ws.cell(row=row, column=col, value=item.get("name", ""))
            col += 1
            ws.cell(row=row, column=col, value=item.get("brand", ""))
            col += 1
            ws.cell(row=row, column=col, value=item.get("model", ""))
            col += 1
            ws.cell(row=row, column=col, value=item.get("specification", ""))
            col += 1
            
            if can_view_internal:
                internal_price = item.get("internal_price", 0) or 0
                ws.cell(row=row, column=col, value=internal_price)
                col += 1
            
            tax_included_price = item.get("tax_included_price", 0)
            quantity = item.get("quantity", 1)
            subtotal = tax_included_price * quantity
            
            ws.cell(row=row, column=col, value=tax_included_price)
            col += 1
            ws.cell(row=row, column=col, value=quantity)
            col += 1
            ws.cell(row=row, column=col, value=subtotal)
        
        # 写入总计行
        totals = calculate_order_totals(items, include_internal=can_view_internal)
        total_row = len(items) + 3
        
        ws.cell(row=total_row, column=1, value="总计").font = Font(bold=True)
        if can_view_internal:
            ws.cell(row=total_row, column=5, value=totals["total_internal_price"]).font = Font(bold=True)
            ws.cell(row=total_row, column=8, value=totals["total_tax_included_price"]).font = Font(bold=True)
        else:
            ws.cell(row=total_row, column=7, value=totals["total_tax_included_price"]).font = Font(bold=True)
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        if can_view_internal:
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 8
            ws.column_dimensions['H'].width = 12
        else:
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 8
            ws.column_dimensions['G'].width = 12
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(
            f"订单导出成功: 订单ID={order_id}, 用户={current_user.username}",
            extra={
                "order_id": order_id,
                "user_id": current_user.id,
                "user_type": current_user.user_type
            }
        )
        
        # 使用urllib.parse.quote对文件名进行编码，避免中文字符编码问题
        # 使用RFC 5987格式支持UTF-8编码的文件名
        filename = f"订单_{order_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        encoded_filename = quote(filename, safe='')
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(
            f"导出订单失败: {e}",
            exc_info=True,
            extra={
                "order_id": order_id,
                "user_id": current_user.id if current_user else None
            }
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="导出订单失败"
        )

